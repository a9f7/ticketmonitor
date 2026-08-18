#!/usr/bin/env python3
# 导出机票监控历史爬取记录为 Excel（多 sheet）
#   Sheet 1 航线明细（最新快照）
#   Sheet 2 价格历史（每条 航线×日期组合，含首/末次抓取与样本数）
#   Sheet 3 抓取健康度（回应"末次抓取集中在最后一天"：分模块统计抓取天数、样本分布、可对比比例）
#   Sheet 4 日本红叶季专项（japan-koyo 模块单独成表，便于核查"航线少"问题）
import json, os, re, datetime, collections
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MODULES = ['gba-summer', 'japan-koyo', 'global-year']
MOD_CN = {'gba-summer': '大湾区寒暑期', 'japan-koyo': '日本红叶季', 'global-year': '全球低价(1年)'}
OUT = os.path.join(ROOT, 'exports', '机票航线记录_导出.xlsx')

code_city = {}  # 代码 -> 城市名

def load_json(p):
    with open(p, encoding='utf-8') as f:
        return json.load(f)

# ---- Sheet 1: 航线明细 ----
wb = Workbook()
ws1 = wb.active
ws1.title = '航线明细'
headers1 = ['模块', '出发城市', '出发代码', '到达城市', '到达代码', '区域', '距离(km)',
            '价格档位', '最低价(¥)', '日历中位(¥)', '日历最高(¥)', '折扣(%)',
            '预算内日期对数', '总日期对数', '选项数', '是否受限',
            '最近最低价去程', '最近最低价回程', '历史样本数', '可对比(>=3样本)', '抓取时间']
ws1.append(headers1)

# 预载历史样本数（按 出发>到达 聚合），便于标注"可对比"
hist_samples = {}   # mod -> (origin>dest) -> 样本数
for mod in MODULES:
    ph = os.path.join(ROOT, 'data', mod, 'price_history.json')
    hist_samples[mod] = collections.defaultdict(int)
    if os.path.exists(ph):
        for k, v in load_json(ph).items():
            oc = k.split('>')[0]; dc = k.split('>')[1].split('@')[0]
            hist_samples[mod][oc + '>' + dc] += len(v.get('at', []))

for mod in MODULES:
    d = load_json(os.path.join(ROOT, 'data', mod, 'flights.json'))
    for o in d.get('origins', []):
        code_city[o['code']] = o.get('city', o['code'])
    for r in d.get('routes', []):
        code_city[r['code']] = r.get('city', r['code'])
        code_city[r['originCode']] = r.get('originCity', r['originCode'])
        cp = (r.get('cheapestPairs') or [])
        od = r['originCode'] + '>' + r['code']
        n = hist_samples[mod].get(od, 0)
        ws1.append([
            MOD_CN.get(mod, mod), r.get('originCity', ''), r.get('originCode', ''),
            r.get('city', ''), r.get('code', ''), r.get('region', '') or r.get('area', ''),
            r.get('distanceKm', ''), r.get('tier', ''), r.get('minPrice', ''),
            r.get('calMedian', ''), r.get('calMax', ''), r.get('discountPct', ''),
            r.get('datePairsInBudget', ''), r.get('totalPairs', ''), r.get('optionCount', ''),
            '是' if r.get('isLimited') else '否',
            cp[0].get('dep', '') if cp else '', cp[0].get('ret', '') if cp else '',
            n, '是' if n >= 3 else '否', d.get('generatedAt', ''),
        ])

# ---- Sheet 2: 价格历史 ----
ws2 = wb.create_sheet('价格历史')
headers2 = ['模块', '出发', '到达', '出发城市', '到达城市', '去程日期', '回程日期',
            '行程天数', '抓取点数', '组合最低', '组合中位', '组合最高', '最新价(¥)', '首次抓取', '末次抓取']
ws2.append(headers2)
key_re = re.compile(r'^([A-Z]{3})>([A-Z]{3})@(\d{4}-\d{2}-\d{2})->(\d{4}-\d{2}-\d{2})$')
for mod in MODULES:
    ph_path = os.path.join(ROOT, 'data', mod, 'price_history.json')
    if not os.path.exists(ph_path):
        continue
    ph = load_json(ph_path)
    for key, rec in ph.items():
        m = key_re.match(key)
        if not m:
            parts = key.split('@')
            if len(parts) != 2:
                continue
            oc, dc = parts[0].split('>')
            dd, rd = parts[1].split('->')
        else:
            oc, dc, dd, rd = m.group(1), m.group(2), m.group(3), m.group(4)
        try:
            days = (datetime.date.fromisoformat(rd) - datetime.date.fromisoformat(dd)).days
        except Exception:
            days = ''
        prices = rec.get('prices', [])
        ats = rec.get('at', [])
        ws2.append([
            MOD_CN.get(mod, mod), oc, dc, code_city.get(oc, ''), code_city.get(dc, ''),
            dd, rd, days, rec.get('count', len(prices)),
            rec.get('min', ''), rec.get('median', ''), rec.get('max', ''),
            prices[-1] if prices else '', ats[0] if ats else '', ats[-1] if ats else '',
        ])

# ---- Sheet 3: 抓取健康度 ----
ws3 = wb.create_sheet('抓取健康度')
headers3 = ['模块', '历史记录数', '去重航线(出发>到达)', '抓取天数', '累计样本数',
            '末次抓取=最新一天占比', '仅1次抓取(不可比)', '可对比航线占比(>=3样本)']
ws3.append(headers3)
for mod in MODULES:
    ph_path = os.path.join(ROOT, 'data', mod, 'price_history.json')
    if not os.path.exists(ph_path):
        continue
    ph = load_json(ph_path)
    last = collections.Counter(); first = collections.Counter()
    only1 = 0; total = len(ph); samples = 0
    od_set = set(); od_samples = collections.defaultdict(int)
    for k, v in ph.items():
        ats = v.get('at', [])
        for a in ats:
            last[a[:10]] += 1
        if len(ats) == 1:
            only1 += 1
        samples += len(ats)
        oc = k.split('>')[0]; dc = k.split('>')[1].split('@')[0]
        od_set.add(oc + '>' + dc); od_samples[oc + '>' + dc] += len(ats)
    # 该模块最新一次抓取日 = 末次抓取中出现最多的日期（近似"最新一天"）
    latest_day = last.most_common(1)[0][0] if last else ''
    last_share = (last[latest_day] / samples * 100) if samples else 0
    ge3 = sum(1 for od in od_set if od_samples[od] >= 3)
    ws3.append([
        MOD_CN.get(mod, mod), total, len(od_set), len(last), samples,
        '%.1f%%' % last_share, only1, '%.1f%%' % (ge3 / len(od_set) * 100 if od_set else 0),
    ])

# ---- Sheet 4: 日本红叶季专项 ----
ws4 = wb.create_sheet('日本红叶季专项')
headers4 = ['出发城市', '到达城市', '到达代码', '区域(地方)', '最低价(¥)', '日历中位(¥)', '价格档位',
            '红叶阶段', '预算内日期对数', '总日期对数', '历史样本数', '可对比(>=3样本)']
ws4.append(headers4)
jp = os.path.join(ROOT, 'data', 'japan-koyo', 'flights.json')
if os.path.exists(jp):
    jd = load_json(jp)
    area_cn = {'hokkaido':'北海道','tohoku':'东北','kanto':'关东','chubu':'中部','kansai':'关西',
               'chugoku':'中国','shikoku':'四国','kyushu':'九州','okinawa':'冲绳'}
    for r in jd.get('routes', []):
        od = r['originCode'] + '>' + r['code']
        n = hist_samples['japan-koyo'].get(od, 0)
        ws4.append([
            r.get('originCity', ''), r.get('city', ''), r.get('code', ''),
            area_cn.get(r.get('area', ''), r.get('area', '')), r.get('minPrice', ''),
            r.get('calMedian', ''), r.get('tier', ''), r.get('koyoLabel', ''),
            r.get('datePairsInBudget', ''), r.get('totalPairs', ''), n,
            '是' if n >= 3 else '否',
        ])

# ---- 样式 ----
hdr_fill = PatternFill('solid', fgColor='2F5496')
hdr_font = Font(bold=True, color='FFFFFF')
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for ws, headers in ((ws1, headers1), (ws2, headers2), (ws3, headers3), (ws4, headers4)):
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = max(10, min(22, len(headers[c - 1]) * 1.7 + 4))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f'已生成: {OUT}')
print(f'  航线明细: {ws1.max_row - 1} 行')
print(f'  价格历史: {ws2.max_row - 1} 行')
print(f'  抓取健康度: {ws3.max_row - 1} 行')
print(f'  日本红叶季专项: {ws4.max_row - 1} 行')
